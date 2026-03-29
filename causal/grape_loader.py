"""
grape_loader.py — Load GRAPE Excel data into GRAPEParticipant model objects.

Usage
-----
    from grape_loader import load_grape_excel, participants_to_dataframe

    participants = load_grape_excel("VF_and_clinical_information_mini.xlsx")
    df = participants_to_dataframe(participants)   # shared causal nodes
    df_full = participants_to_dataframe(participants, full=True)

Sheet layout
------------
  Baseline  : row 1 = group headers, row 2 = sub-headers, row 3+ = data
  Follow-Up : row 1 = group headers, row 2 = sub-column indices, row 3+ = data
"""

from __future__ import annotations

from pathlib import Path
from typing import Union

import openpyxl

from base_models import BaseDemographics
from grape_models import (
    BaselineOcularData,
    Exam,
    ExamRecord,
    GRAPEDiagnosis,
    GRAPEEyeExam,
    GRAPEFundusImage,
    GRAPEParticipant,
    ProgressionStatus,
    RNFLMeasurement,
    VisualField,
)

# ── Column indices (0-based) ──────────────────────────────────────────────────

BL_SUBJECT      = 0
BL_LATERALITY   = 1
BL_AGE          = 2
BL_GENDER       = 3
BL_IOP          = 4
BL_CCT          = 5
BL_TOTAL_VISITS = 6
BL_PLR2         = 7
BL_PLR3         = 8
BL_MD_FLAG      = 9
BL_GLAUCOMA_CAT = 10
BL_RNFL_MEAN    = 11
BL_RNFL_S       = 12
BL_RNFL_N       = 13
BL_RNFL_I       = 14
BL_RNFL_T       = 15
BL_CFP          = 16
BL_CAMERA       = 17
BL_RESOLUTION   = 18
BL_VF_START     = 19   # columns 19–79 → VF points 0–60

FU_SUBJECT      = 0
FU_LATERALITY   = 1
FU_VISIT_NUM    = 2
FU_INTERVAL     = 3
FU_IOP          = 4
FU_CFP          = 5
FU_CAMERA       = 6
FU_RESOLUTION   = 7
FU_VF_START     = 8    # columns 8–68 → VF points 0–60

VF_N = 61


# ── Helpers ───────────────────────────────────────────────────────────────────

def _opt_float(val) -> float | None:
    try:
        return float(val) if val is not None else None
    except (TypeError, ValueError):
        return None


def _opt_int(val) -> int | None:
    try:
        return int(val) if val is not None else None
    except (TypeError, ValueError):
        return None


def _parse_vf(row: tuple, start_col: int) -> VisualField:
    raw = list(row[start_col: start_col + VF_N])
    return VisualField.from_raw(raw)


def _parse_fundus(cfp, camera, resolution, laterality=None) -> GRAPEFundusImage:
    return GRAPEFundusImage.from_raw(
        filename=str(cfp) if cfp else None,
        camera=str(camera) if camera else None,
        resolution_str=str(resolution) if resolution else None,
        laterality=laterality,
    )


# ── Baseline loader ───────────────────────────────────────────────────────────

def _load_baseline(ws) -> dict[int, GRAPEParticipant]:
    participants: dict[int, GRAPEParticipant] = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        subject_id = _opt_int(row[BL_SUBJECT])
        if subject_id is None:
            continue

        laterality = str(row[BL_LATERALITY]).strip() if row[BL_LATERALITY] else None

        if subject_id not in participants:
            cat = str(row[BL_GLAUCOMA_CAT]).strip() if row[BL_GLAUCOMA_CAT] else ""
            dx = GRAPEDiagnosis.from_category(cat)
            dx.progression = ProgressionStatus(
                plr2=_opt_int(row[BL_PLR2]),
                plr3=_opt_int(row[BL_PLR3]),
                md=_opt_int(row[BL_MD_FLAG]),
            )
            participants[subject_id] = GRAPEParticipant(
                subject_id=subject_id,
                demographics=BaseDemographics(
                    age=_opt_int(row[BL_AGE]),
                    gender=str(row[BL_GENDER]).strip() if row[BL_GENDER] else None,
                ),
                diagnosis=dx,
                total_visits=_opt_int(row[BL_TOTAL_VISITS]),
            )

        p = participants[subject_id]

        if laterality:
            p.baseline_ocular[laterality] = BaselineOcularData(
                laterality=laterality,
                cct=_opt_float(row[BL_CCT]),
                rnfl=RNFLMeasurement(
                    mean=_opt_float(row[BL_RNFL_MEAN]),
                    superior=_opt_float(row[BL_RNFL_S]),
                    nasal=_opt_float(row[BL_RNFL_N]),
                    inferior=_opt_float(row[BL_RNFL_I]),
                    temporal=_opt_float(row[BL_RNFL_T]),
                ),
            )

    return participants


# ── Follow-Up loader ──────────────────────────────────────────────────────────

def _load_followup(ws, participants: dict[int, GRAPEParticipant]) -> None:
    exam_map: dict[int, dict[int, Exam]] = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        subject_id = _opt_int(row[FU_SUBJECT])
        if subject_id is None:
            continue

        laterality  = str(row[FU_LATERALITY]).strip() if row[FU_LATERALITY] else None
        visit_num   = _opt_int(row[FU_VISIT_NUM]) or 1
        interval    = _opt_float(row[FU_INTERVAL]) or 0.0
        is_baseline = (visit_num == 1)

        vf = _parse_vf(row, FU_VF_START)
        eye_exam = GRAPEEyeExam(
            laterality=laterality,
            iop=_opt_float(row[FU_IOP]),
            visual_field=vf,
            vf_md=vf.mean_sensitivity,   # populate shared DAG node explicitly
            fundus=_parse_fundus(
                row[FU_CFP], row[FU_CAMERA], row[FU_RESOLUTION], laterality
            ),
        )

        if subject_id not in exam_map:
            exam_map[subject_id] = {}

        if visit_num not in exam_map[subject_id]:
            exam_map[subject_id][visit_num] = Exam(
                visit_number=visit_num,
                is_baseline=is_baseline,
                interval_years=interval,
            )

        if laterality:
            exam_map[subject_id][visit_num].eyes[laterality] = eye_exam

    for subject_id, visits in exam_map.items():
        if subject_id not in participants:
            participants[subject_id] = GRAPEParticipant(subject_id=subject_id)

        p = participants[subject_id]
        sorted_exams = [visits[vn] for vn in sorted(visits)]
        p.exam_record = ExamRecord(
            total_exams=len(sorted_exams),
            exams=sorted_exams,
        )
        p.compute_asymmetry()


# ── Public API ────────────────────────────────────────────────────────────────

def load_grape_excel(path: Union[str, Path]) -> dict[int, GRAPEParticipant]:
    """
    Load the GRAPE Excel file and return dict[subject_id, GRAPEParticipant].
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"GRAPE Excel file not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    if "Baseline" not in wb.sheetnames:
        raise ValueError("Expected a 'Baseline' sheet in the workbook.")

    print("[grape_loader] Loading Baseline sheet …")
    participants = _load_baseline(wb["Baseline"])
    print(f"[grape_loader] Loaded {len(participants)} participants from Baseline.")

    if "Follow-Up" in wb.sheetnames:
        print("[grape_loader] Loading Follow-Up sheet …")
        _load_followup(wb["Follow-Up"], participants)
        n = sum(1 for p in participants.values() if p.exam_record.total_exams > 0)
        print(f"[grape_loader] Exam records attached for {n} participants.")
    else:
        print("[grape_loader] No 'Follow-Up' sheet; exam_record will be empty.")

    return participants


def participants_to_dataframe(
    participants: dict[int, GRAPEParticipant],
    full: bool = False,
):
    """
    Build a DataFrame from loaded GRAPE participants.

    full=False (default): shared causal DAG nodes only — safe to concat with
                          PAPILA output from papila_loader.participants_to_dataframe().
    full=True:            all fields including GRAPE-specific (RNFL, full VF, etc.)
    """
    import pandas as pd
    rows = []
    for p in participants.values():
        rows.append(p.to_flat_dict() if full else p.causal_node_dict())
    return pd.DataFrame(rows)
